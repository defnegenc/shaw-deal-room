from dataclasses import dataclass
from datetime import date
import re
from pathlib import Path


@dataclass(frozen=True)
class ExtractedFact:
    field_name: str
    value_text: str | None
    value_numeric: float | None
    unit: str | None
    currency: str | None
    as_of_date: date | None
    evidence: str
    confidence_score: float
    extraction_method: str = "regex"
    source_url: str | None = None


MONEY_FIELDS = {
    "arr": [r"\bARR:\s*\$?([\d.]+)\s*([MK])?\b"],
    "monthly_burn": [r"\bMonthly burn:\s*\$?([\d.]+)\s*([MK])?\b"],
    "target_raise": [r"\bTarget raise:\s*\$?([\d.]+)\s*([MK])?\b"],
    "pre_money_valuation": [r"\bPre-money valuation:\s*\$?([\d.]+)\s*([MK])?\b"],
    "post_money_valuation": [r"\bPost-money valuation:\s*\$?([\d.]+)\s*([MK])?\b"],
    "investment_amount": [r"\bInvestment amount:\s*\$?([\d.]+)\s*([MK])?\b"],
}

NUMBER_FIELDS = {
    "headcount": [r"\bTeam size:\s*(\d+)\s*employees\b", r"\b(\d+)\s*employees\b"],
    "revenue_growth_pct": [r"\bRevenue growth:\s*([\d.]+)%"],
    "gross_margin_pct": [r"\bGross margin:\s*([\d.]+)%"],
    "runway_months": [r"\bRunway:\s*([\d.]+)\s*months\b"],
}

TEXT_FIELDS = {
    "headquarters": [r"\bHeadquarters:\s*(.+)"],
    "round": [r"\b(?:Current round|Round):\s*(.+)"],
    "lead_investor": [r"\bLead investor:\s*(.+)"],
}


def parse_document_type(text: str) -> str:
    match = re.search(r"Document type:\s*(.+)", text, re.IGNORECASE)
    return normalize_document_type(match.group(1)) if match else "unknown"


def normalize_document_type(value: str | None) -> str:
    if not value:
        return "unknown"
    label = re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
    aliases = {
        "email_screenshot": "email_screenshot",
        "screenshot": "email_screenshot",
        "founder_email": "email",
        "email": "email",
        "handwritten_note": "handwritten_note",
        "handwritten_notes": "handwritten_note",
        "excel": "spreadsheet",
        "excel_sheet": "spreadsheet",
        "xlsx": "spreadsheet",
        "csv": "spreadsheet",
        "spreadsheet": "spreadsheet",
        "pitch_book": "pitch_deck",
        "pitch_deck": "pitch_deck",
        "term_sheet": "term_sheet",
    }
    return aliases.get(label, label or "unknown")


def parse_as_of_date(text: str) -> date | None:
    match = re.search(r"As of:\s*(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
    if not match:
        return None
    return date.fromisoformat(match.group(1))


def read_text(path: str | Path) -> str:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".heic"}:
        return (
            "Document type: image_upload\n"
            f"Filename: {file_path.name}\n"
            "Image text extraction is not implemented in this MVP. "
            "Production parser lane: OCR or multimodal LLM extraction with citation bounding boxes."
        )
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("PDF parsing requires pypdf. Install requirements.txt.") from exc
        reader = PdfReader(str(file_path))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel parsing requires openpyxl. Install requirements.txt.") from exc
        workbook = load_workbook(file_path, data_only=True)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows)
    if suffix == ".csv":
        return file_path.read_text(encoding="utf-8")
    return file_path.read_text(encoding="utf-8")


def chunk_text(text: str, max_chars: int = 900) -> list[str]:
    paragraphs = [part.strip() for part in text.split("\n\n") if part.strip()]
    chunks: list[str] = []
    current = ""
    for paragraph in paragraphs:
        if len(current) + len(paragraph) + 2 > max_chars and current:
            chunks.append(current)
            current = paragraph
        else:
            current = f"{current}\n\n{paragraph}".strip()
    if current:
        chunks.append(current)
    return chunks


def extract_facts(text: str) -> list[ExtractedFact]:
    as_of_date = parse_as_of_date(text)
    facts: list[ExtractedFact] = []

    for field_name, patterns in MONEY_FIELDS.items():
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value = _scaled_number(match.group(1), match.group(2))
                facts.append(
                    ExtractedFact(
                        field_name=field_name,
                        value_text=f"${match.group(1)}{match.group(2) or ''}",
                        value_numeric=value,
                        unit=None,
                        currency="USD",
                        as_of_date=as_of_date,
                        evidence=_line_for_match(text, match.start()),
                        confidence_score=0.95,
                    )
                )
                break

    for field_name, patterns in NUMBER_FIELDS.items():
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                facts.append(
                    ExtractedFact(
                        field_name=field_name,
                        value_text=match.group(1),
                        value_numeric=float(match.group(1)),
                        unit=_unit_for_number_field(field_name),
                        currency=None,
                        as_of_date=as_of_date,
                        evidence=_line_for_match(text, match.start()),
                        confidence_score=0.92 if ":" in _line_for_match(text, match.start()) else 0.78,
                    )
                )
                break

    for field_name, patterns in TEXT_FIELDS.items():
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                facts.append(
                    ExtractedFact(
                        field_name=field_name,
                        value_text=match.group(1).strip(),
                        value_numeric=None,
                        unit=None,
                        currency=None,
                        as_of_date=as_of_date,
                        evidence=_line_for_match(text, match.start()),
                        confidence_score=0.9,
                    )
                )
                break

    return facts


def _scaled_number(value: str, suffix: str | None) -> float:
    number = float(value)
    if suffix and suffix.upper() == "M":
        return number * 1_000_000
    if suffix and suffix.upper() == "K":
        return number * 1_000
    return number


def _unit_for_number_field(field_name: str) -> str | None:
    return {
        "headcount": "employees",
        "revenue_growth_pct": "percent",
        "gross_margin_pct": "percent",
        "runway_months": "months",
    }.get(field_name)


def _line_for_match(text: str, index: int) -> str:
    start = text.rfind("\n", 0, index) + 1
    end = text.find("\n", index)
    if end == -1:
        end = len(text)
    return text[start:end].strip()
